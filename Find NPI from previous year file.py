import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, filedialog

# Function to open a file dialog to select a file
def select_file(title="Select a file"):
    root = Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_path

# Function to open a file dialog to choose where to save the file
def save_file():
    root = Tk()
    root.withdraw()  # Hide the main window
    save_path = filedialog.asksaveasfilename(title="Save As", defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")])
    return save_path

# Select the 'to edit' file and 'to search' file
to_edit_file = select_file("Select the Excel file you want to edit")
if not to_edit_file:
    print("No file selected for editing. Exiting...")
    exit()

to_search_file = select_file("Select the Excel file from which you want to search")
if not to_search_file:
    print("No file selected for searching. Exiting...")
    exit()

# Read the Excel files
to_edit_df = pd.read_excel(to_edit_file)
to_search_df = pd.read_excel(to_search_file)

# Add new column for 'new npi'
to_edit_df['new npi'] = ''

# Convert columns to lowercase to avoid case sensitivity issues
to_edit_df['HCP FIRST NAME'] = to_edit_df['HCP FIRST NAME'].str.lower()
to_edit_df['HCP LAST NAME'] = to_edit_df['HCP LAST NAME'].str.lower()
to_edit_df['ZIP CODE'] = to_edit_df['ZIP CODE'].astype(str).str.zfill(5)

to_search_df['SRC RECIPIENT FIRST NAME'] = to_search_df['SRC RECIPIENT FIRST NAME'].str.lower()
to_search_df['SRC RECIPIENT LAST NAME'] = to_search_df['SRC RECIPIENT LAST NAME'].str.lower()
to_search_df['SRC RECIPIENT ZIP CODE'] = to_search_df['SRC RECIPIENT ZIP CODE'].astype(str).str.zfill(5)

# Define the colors for highlighting
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Iterate through 'to edit' file
for i, row in to_edit_df.iterrows():
    first_name = row['HCP FIRST NAME']
    last_name = row['HCP LAST NAME']
    zip_code = row['ZIP CODE']

    # Search for matching rows in 'to search' file
    match = to_search_df[
        (to_search_df['SRC RECIPIENT FIRST NAME'] == first_name) & 
        (to_search_df['SRC RECIPIENT LAST NAME'] == last_name) & 
        (to_search_df['SRC RECIPIENT ZIP CODE'] == zip_code)
    ]
    
    if not match.empty:
        npi_value = match.iloc[0]['QD UPDATED NPI']
        if pd.notna(npi_value) and npi_value != '':
            to_edit_df.at[i, 'new npi'] = npi_value
        else:
            to_edit_df.at[i, 'new npi'] = ''
    else:
        name_match = to_search_df[
            (to_search_df['SRC RECIPIENT FIRST NAME'] == first_name) & 
            (to_search_df['SRC RECIPIENT LAST NAME'] == last_name)
        ]
        if not name_match.empty:
            to_edit_df.at[i, 'new npi'] = ''

# Save the DataFrame with the new column to a temporary file
temp_file_path = 'temp_file.xlsx'
to_edit_df.to_excel(temp_file_path, index=False)

# Load the temporary workbook to apply formatting
wb = load_workbook(temp_file_path)
ws = wb.active

# Reapply formatting
for i, row in to_edit_df.iterrows():
    first_name = row['HCP FIRST NAME']
    last_name = row['HCP LAST NAME']
    zip_code = row['ZIP CODE']
    new_npi = row['new npi']

    # Determine the row's fill color
    if new_npi:
        fill = green_fill
    else:
        match = to_search_df[
            (to_search_df['SRC RECIPIENT FIRST NAME'] == first_name) & 
            (to_search_df['SRC RECIPIENT LAST NAME'] == last_name) & 
            (to_search_df['SRC RECIPIENT ZIP CODE'] == zip_code)
        ]
        if not match.empty:
            fill = yellow_fill
        else:
            name_match = to_search_df[
                (to_search_df['SRC RECIPIENT FIRST NAME'] == first_name) & 
                (to_search_df['SRC RECIPIENT LAST NAME'] == last_name)
            ]
            if not name_match.empty:
                fill = blue_fill
            else:
                fill = None

    # Apply the fill color to the row
    if fill:
        for cell in ws[i+2]:  # i+2 because pandas is 0-indexed and Excel is 1-indexed with a header row
            cell.fill = fill

# Ask the user where to save the updated file
save_path = save_file()
if save_path:
    wb.save(save_path)  # Save the workbook with formatting
    print(f"Results saved to {save_path}")

# Remove the temporary file
import os
os.remove(temp_file_path)
